from __future__ import annotations

import io
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .database import Database
from .forms import FieldType, get_form_definition, validate_responses
from .models import Inspection, InspectionNote, InspectionType, Truck, User, UserRole

PHOTO_MIN = 4
PHOTO_MAX = 10
NOTE_WINDOW_HOURS = 24


@dataclass
class InspectionService:
    database: Database

    def list_forms(self) -> dict[str, list[dict[str, object]]]:
        return {
            inspection_type.value: [
                {
                    "id": field.id,
                    "label": field.label,
                    "field_type": field.field_type.value,
                    "required": field.required,
                }
                for field in get_form_definition(inspection_type)
            ]
            for inspection_type in InspectionType
        }

    def create_inspection(
        self,
        *,
        inspection_type: InspectionType,
        truck: Truck,
        ranger: User,
        responses: dict,
        photo_urls: Iterable[str],
        video_url: Optional[str],
        escalate_visibility: bool = False,
    ) -> Inspection:
        if ranger.role not in {UserRole.RANGER, UserRole.SUPERVISOR}:
            raise PermissionError("Only rangers and supervisors may create inspections")
        if not truck.active:
            raise ValueError("Truck is not active")
        photo_list = list(photo_urls)
        if inspection_type is not InspectionType.RETURN:
            if len(photo_list) < PHOTO_MIN or len(photo_list) > PHOTO_MAX:
                raise ValueError("Between 4 and 10 photos are required")
        else:
            if len(photo_list) > PHOTO_MAX:
                raise ValueError("At most 10 photos are allowed")
        validated = validate_responses(inspection_type, responses)
        return self.database.add_inspection(
            inspection_type=inspection_type,
            truck_id=truck.id,
            ranger_id=ranger.id,
            escalate_visibility=escalate_visibility,
            responses=validated,
            photo_urls=photo_list,
            video_url=video_url,
        )

    def list_inspections(
        self,
        *,
        requester: User,
        truck: Optional[Truck] = None,
        ranger: Optional[User] = None,
    ) -> List[Inspection]:
        truck_id = truck.id if truck else None
        ranger_filter = ranger.id if ranger else None
        if requester.role == UserRole.RANGER:
            ranger_filter = requester.id
        inspections = list(self.database.list_inspections(truck_id=truck_id, ranger_id=ranger_filter))
        return inspections

    def get_inspection(self, *, requester: User, inspection_id: int) -> Inspection:
        inspection = self.database.get_inspection(inspection_id)
        if not inspection:
            raise LookupError("Inspection not found")
        if requester.role == UserRole.RANGER and inspection.ranger_id != requester.id:
            raise PermissionError("Rangers may only view their own inspections")
        return inspection

    def add_note(self, *, requester: User, inspection: Inspection, content: str) -> InspectionNote:
        if requester.role == UserRole.RANGER and inspection.ranger_id != requester.id:
            raise PermissionError("Cannot annotate another ranger's inspection")
        if datetime.utcnow() - inspection.created_at > timedelta(hours=NOTE_WINDOW_HOURS):
            raise ValueError("Notes can only be added within 24 hours of submission")
        note = self.database.add_note(inspection_id=inspection.id, author_id=requester.id, content=content)
        self.database.update_inspection_timestamp(inspection.id, note.created_at)
        return note

    def list_notes(self, inspection_id: int) -> List[InspectionNote]:
        return list(self.database.list_notes(inspection_id))

    def export_inspections_workbook(self, *, generated_by: User) -> tuple[str, bytes]:
        inspections = list(self.database.list_inspections())
        trucks = {
            truck_id: self.database.get_truck(truck_id)
            for truck_id in {inspection.truck_id for inspection in inspections}
        }
        rangers = {
            ranger_id: self.database.get_user(ranger_id)
            for ranger_id in {inspection.ranger_id for inspection in inspections}
        }
        notes_count: dict[int, int] = {}
        for inspection in inspections:
            notes_count[inspection.id] = sum(1 for _ in self.database.list_notes(inspection.id))

        workbook = Workbook()
        summary_ws = workbook.active
        summary_ws.title = "Summary"

        now = datetime.utcnow()
        title_font = Font(size=16, bold=True, color="24512C")
        header_font = Font(bold=True, color="1F2A24")
        muted_font = Font(color="5B6657")

        summary_ws["A1"] = "Inspection program snapshot"
        summary_ws["A1"].font = title_font
        summary_ws.merge_cells("A1:E1")
        summary_ws["A2"] = f"Generated for {generated_by.name}"
        summary_ws["A2"].font = muted_font
        summary_ws.merge_cells("A2:E2")
        summary_ws["A3"] = now.strftime("Created %Y-%m-%d %H:%M UTC")
        summary_ws["A3"].font = muted_font
        summary_ws.merge_cells("A3:E3")

        total = len(inspections)
        escalated = sum(1 for inspection in inspections if inspection.escalate_visibility)
        last_inspection = max((inspection.created_at for inspection in inspections), default=None)
        unique_trucks = len({inspection.truck_id for inspection in inspections})
        unique_rangers = len({inspection.ranger_id for inspection in inspections})
        average_photos = (
            round(sum(len(inspection.photo_urls) for inspection in inspections) / total, 1)
            if total
            else 0.0
        )

        summary_ws["A5"], summary_ws["B5"] = "Metric", "Value"
        summary_ws["A5"].font = header_font
        summary_ws["B5"].font = header_font

        metrics = [
            ("Total inspections", total),
            ("Escalated inspections", escalated),
            ("Latest inspection", last_inspection.strftime("%Y-%m-%d %H:%M") if last_inspection else "—"),
            ("Unique trucks", unique_trucks),
            ("Unique rangers", unique_rangers),
            ("Avg. photos per inspection", average_photos),
        ]
        for index, (label, value) in enumerate(metrics, start=6):
            summary_ws.cell(row=index, column=1, value=label)
            summary_ws.cell(row=index, column=2, value=value)

        type_counts = Counter(inspection.inspection_type for inspection in inspections)
        summary_ws["A13"], summary_ws["B13"] = "Inspection type", "Count"
        summary_ws["A13"].font = header_font
        summary_ws["B13"].font = header_font
        row_pointer = 14
        for inspection_type in InspectionType:
            summary_ws.cell(row=row_pointer, column=1, value=inspection_type.value.title())
            summary_ws.cell(row=row_pointer, column=2, value=type_counts.get(inspection_type, 0))
            row_pointer += 1

        ranger_counts = Counter(inspection.ranger_id for inspection in inspections)
        if ranger_counts:
            summary_ws["D5"] = "Most active rangers"
            summary_ws["D5"].font = header_font
            summary_ws["E5"] = "Completed"
            summary_ws["E5"].font = header_font
            for offset, (ranger_id, count) in enumerate(ranger_counts.most_common(3), start=6):
                ranger = rangers.get(ranger_id)
                label = ranger.name if ranger else f"Ranger {ranger_id}"
                summary_ws.cell(row=offset, column=4, value=label)
                summary_ws.cell(row=offset, column=5, value=count)

        for column, width in [(1, 26), (2, 22), (4, 28), (5, 12)]:
            summary_ws.column_dimensions[get_column_letter(column)].width = width

        detail_ws = workbook.create_sheet("Inspections")
        detail_headers = [
            "Inspection #",
            "Submitted (UTC)",
            "Type",
            "Truck",
            "Ranger",
            "Escalated",
            "Odometer",
            "Fuel level (%)",
            "Photos",
            "Video",
            "Notes",
            "Attention items",
        ]
        detail_ws.append(detail_headers)
        for cell in detail_ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for inspection in inspections:
            truck = trucks.get(inspection.truck_id)
            truck_label = truck.identifier if truck else f"Truck {inspection.truck_id}"
            ranger = rangers.get(inspection.ranger_id)
            ranger_label = ranger.name if ranger else f"Ranger {inspection.ranger_id}"
            responses = inspection.responses
            mileage = responses.get("odometer_miles")
            fuel_level = responses.get("fuel_level")
            attention_items: list[str] = []
            for field in get_form_definition(inspection.inspection_type):
                value = responses.get(field.id)
                if field.field_type is FieldType.BOOLEAN:
                    if field.id == "fluid_leak_detected" and value:
                        attention_items.append(field.label)
                    elif value is False:
                        attention_items.append(field.label)
                elif field.field_type is FieldType.TEXT and field.id in {"notes", "return_notes"} and value:
                    attention_items.append(f"{field.label}: {value}")
            attention_text = ", ".join(attention_items)

            row = [
                inspection.id,
                inspection.created_at.strftime("%Y-%m-%d %H:%M"),
                inspection.inspection_type.value.title(),
                truck_label,
                ranger_label,
                "Yes" if inspection.escalate_visibility else "No",
                mileage if mileage is not None else "",
                fuel_level if fuel_level is not None else "",
                len(inspection.photo_urls),
                "Yes" if inspection.video_url else "No",
                notes_count.get(inspection.id, 0),
                attention_text,
            ]
            detail_ws.append(row)

            if inspection.escalate_visibility:
                highlight = PatternFill(start_color="FFF4E5", end_color="FFF4E5", fill_type="solid")
                for cell in detail_ws[detail_ws.max_row]:
                    cell.fill = highlight

        detail_ws.auto_filter.ref = detail_ws.dimensions
        detail_ws.freeze_panes = "A2"

        for column_index in range(1, len(detail_headers) + 1):
            column_letter = get_column_letter(column_index)
            max_length = max(
                (len(str(detail_ws.cell(row=row, column=column_index).value or "")) for row in range(1, detail_ws.max_row + 1)),
                default=10,
            )
            detail_ws.column_dimensions[column_letter].width = min(max(12, max_length + 2), 42)

        timestamp = now.strftime("%Y%m%d-%H%M%S")
        filename = f"inspection-export-{timestamp}.xlsx"
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return filename, buffer.getvalue()

    def personnel_metrics(self) -> list[dict[str, object]]:
        metrics: list[dict[str, object]] = []
        personnel = self.database.list_users_by_roles([UserRole.RANGER, UserRole.SUPERVISOR])
        assignments = list(self.database.list_assignments())
        completed_assignments = [assignment for assignment in assignments if assignment.returned_at is not None]
        for person in personnel:
            user_assignments = [assignment for assignment in completed_assignments if assignment.ranger_id == person.id]
            if user_assignments:
                most_recent = max(assignment.returned_at for assignment in user_assignments if assignment.returned_at is not None)
            else:
                most_recent = None
            metrics.append(
                {
                    "user": person,
                    "role": person.role,
                    "inspections_completed": len(user_assignments),
                    "most_recent_inspection": most_recent,
                }
            )
        return metrics

    def dashboard(self) -> dict[str, object]:
        inspections = list(self.database.list_inspections())
        assignments = list(self.database.list_assignments())
        closed_assignments = [assignment for assignment in assignments if assignment.returned_at is not None]
        escalated = sum(1 for insp in inspections if insp.escalate_visibility)
        return {
            "total_inspections": len(closed_assignments),
            "escalated_inspections": escalated,
            "personnel_metrics": self.personnel_metrics(),
        }
