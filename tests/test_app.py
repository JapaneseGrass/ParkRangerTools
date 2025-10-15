from __future__ import annotations

import base64
import inspect
import io
from datetime import datetime, timedelta
from pathlib import Path

import pytest

from openpyxl import load_workbook

from backend.app import TruckInspectionApp
from backend.app.auth import TOKEN_EXPIRY_MINUTES


def test_dataclasses_do_not_use_slots() -> None:
    from backend.app import app as app_module
    from backend.app import auth as auth_module
    from backend.app import forms as forms_module
    from backend.app import inspections as inspections_module
    from backend.app import models as models_module

    modules = [app_module, auth_module, forms_module, inspections_module, models_module]
    dataclass_params = []
    for module in modules:
        for _, obj in inspect.getmembers(module, inspect.isclass):
            params = getattr(obj, "__dataclass_params__", None)
            if params is not None:
                dataclass_params.append(params)

    assert dataclass_params, "Expected to discover dataclasses in backend modules"
    assert all(
        not getattr(params, "slots", False) for params in dataclass_params
    ), "Dataclasses must not request slots for Python 3.9 compatibility"
from backend.app.models import InspectionType, User, UserRole

_SAMPLE_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAvoB9pWcVYoAAAAASUVORK5CYII="
)


def _write_sample_photo(path: Path) -> None:
    path.write_bytes(_SAMPLE_PNG)


def _quick_responses() -> dict[str, object]:
    return {
        "exterior_clean": True,
        "interior_clean": True,
        "seatbelts_functioning": True,
        "tire_inflation": True,
        "fuel_level": "75",
        "odometer_miles": 1200,
    }


def _detailed_responses() -> dict[str, object]:
    responses = _quick_responses()
    responses.update(
        {
            "engine_oil_ok": True,
            "fan_belts_ok": True,
            "coolant_level_ok": True,
            "washer_fluid_ok": True,
            "wipers_ok": True,
            "tire_tread_ok": True,
            "headlights_ok": True,
            "turn_signals_ok": True,
            "brake_lights_ok": True,
            "reverse_lights_ok": True,
            "fluid_leak_detected": False,
            "mirrors_ok": True,
            "emergency_system_ok": True,
        }
    )
    return responses


def _photos() -> list[str]:
    return [
        "photo1.jpg",
        "photo2.jpg",
        "photo3.jpg",
        "photo4.jpg",
    ]


def test_authentication_success(seeded_app: TruckInspectionApp) -> None:
    token = seeded_app.auth.authenticate("ranger@email.com", "password")
    assert token is not None
    delta = token.expires_at - token.created_at
    assert abs(delta - timedelta(minutes=TOKEN_EXPIRY_MINUTES)) < timedelta(seconds=1)


def test_authentication_failure(seeded_app: TruckInspectionApp) -> None:
    token = seeded_app.auth.authenticate("ranger@email.com", "wrongpass")
    assert token is None


def test_ranger_submits_quick_inspection(seeded_app: TruckInspectionApp, ranger: User, truck) -> None:
    inspection = seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.QUICK,
        responses=_quick_responses(),
        photo_urls=_photos(),
    )
    assert inspection.truck_id == truck.id
    assert inspection.ranger_id == ranger.id
    assert inspection.responses["fuel_level"] == "75"


def test_photo_validation(seeded_app: TruckInspectionApp, ranger: User, truck) -> None:
    with pytest.raises(ValueError):
        seeded_app.submit_inspection(
            user=ranger,
            truck=truck,
            inspection_type=InspectionType.QUICK,
            responses=_quick_responses(),
            photo_urls=_photos()[:3],
        )


def test_notes_within_window(seeded_app: TruckInspectionApp, ranger: User, truck) -> None:
    inspection = seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.DETAILED,
        responses=_detailed_responses(),
        photo_urls=_photos(),
    )
    note = seeded_app.add_note(requester=ranger, inspection=inspection, content="Follow-up")
    assert note.content == "Follow-up"


def test_notes_after_window_fails(seeded_app: TruckInspectionApp, ranger: User, truck) -> None:
    inspection = seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.QUICK,
        responses=_quick_responses(),
        photo_urls=_photos(),
    )
    past = datetime.utcnow() - timedelta(hours=25)
    with seeded_app.database.session() as conn:
        conn.execute(
            "UPDATE inspections SET created_at = ?, updated_at = ? WHERE id = ?",
            (
                past.strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                past.strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
                inspection.id,
            ),
        )
    inspection = seeded_app.database.get_inspection(inspection.id)
    assert inspection is not None
    with pytest.raises(ValueError):
        seeded_app.add_note(requester=ranger, inspection=inspection, content="Too late")


def test_dashboard_metrics(seeded_app: TruckInspectionApp, ranger: User, supervisor: User, truck) -> None:
    seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.QUICK,
        responses=_quick_responses(),
        photo_urls=_photos(),
    )
    seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.QUICK,
        responses=_quick_responses(),
        photo_urls=_photos(),
        escalate_visibility=True,
    )
    # Perform checkout/return to contribute to compliance totals
    start_inspection = seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.QUICK,
        responses=_quick_responses(),
        photo_urls=_photos(),
    )
    assignment = seeded_app.checkout_truck(ranger=ranger, truck=truck, inspection=start_inspection)
    end_inspection = seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.RETURN,
        responses={"odometer_miles": 2000, "return_notes": "Done"},
        photo_urls=[],
    )
    seeded_app.return_truck(assignment_id=assignment.id, ranger=ranger, inspection=end_inspection)

    dashboard = seeded_app.dashboard(supervisor=supervisor)
    assert dashboard["total_inspections"] == 1
    assert dashboard["escalated_inspections"] == 1
    metrics = dashboard["personnel_metrics"]
    ranger_entry = next(entry for entry in metrics if entry["user"].id == ranger.id)
    supervisor_entry = next(entry for entry in metrics if entry["user"].id == supervisor.id)
    assert ranger_entry["inspections_completed"] == 1
    assert supervisor_entry["inspections_completed"] == 0


def test_supervisor_excel_export(
    seeded_app: TruckInspectionApp,
    supervisor: User,
    ranger: User,
    truck,
    tmp_path: Path,
) -> None:
    detailed = _detailed_responses()
    detailed["seatbelts_functioning"] = False
    detailed["fluid_leak_detected"] = True
    detailed["notes"] = "Seatbelt frayed on driver side"
    photo_dir = tmp_path / "photos"
    photo_dir.mkdir()
    photo_urls: list[str] = []
    for index in range(1, 5):
        path = photo_dir / f"inspection-{index}.png"
        _write_sample_photo(path)
        photo_urls.append(str(path))
    inspection = seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.DETAILED,
        responses=detailed,
        photo_urls=photo_urls,
        escalate_visibility=True,
    )
    seeded_app.add_note(
        requester=supervisor,
        inspection=inspection,
        content="Schedule maintenance review",
    )

    return_inspection = seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.RETURN,
        responses={"odometer_miles": 1300, "return_notes": "Left keys with dispatch"},
        photo_urls=[],
    )

    filename, payload = seeded_app.export_inspections(
        supervisor=supervisor,
        photo_resolver=lambda url: Path(url) if Path(url).exists() else None,
    )
    assert filename.startswith("inspection-export-")
    assert filename.endswith(".xlsx")
    workbook = load_workbook(io.BytesIO(payload))
    assert set(workbook.sheetnames) >= {"Summary", "Inspections"}

    summary = workbook["Summary"]
    metrics: dict[str, object] = {}
    for row in range(6, 12):
        label = summary.cell(row=row, column=1).value
        value = summary.cell(row=row, column=2).value
        if label:
            metrics[label] = value
    assert metrics.get("Total inspections") == 2
    assert metrics.get("Escalated inspections") == 1

    detail = workbook["Inspections"]
    rows: dict[int, dict[str, object]] = {}
    for index in range(2, detail.max_row + 1):
        inspection_id = detail.cell(row=index, column=1).value
        if inspection_id is None:
            continue
        rows[inspection_id] = {
            "escalated": detail.cell(row=index, column=6).value,
            "notes": detail.cell(row=index, column=11).value,
            "attention": detail.cell(row=index, column=12).value or "",
        }

    assert inspection.id in rows
    assert rows[inspection.id]["escalated"] == "Yes"
    assert rows[inspection.id]["notes"] == 1
    assert "Seatbelt" in rows[inspection.id]["attention"]
    assert "Fluid leaking" in rows[inspection.id]["attention"]

    assert return_inspection.id in rows
    assert rows[return_inspection.id]["escalated"] == "No"
    assert rows[return_inspection.id]["notes"] == 0
    assert "Return notes" in rows[return_inspection.id]["attention"]

    responses_ws = workbook["Responses"]
    response_pairs = {
        (responses_ws.cell(row=row, column=1).value, responses_ws.cell(row=row, column=2).value)
        for row in range(2, responses_ws.max_row + 1)
    }
    assert (inspection.id, "Seatbelts functioning properly") in response_pairs
    assert (inspection.id, "Notes section") in response_pairs

    photos_ws = workbook["Photos"]
    sources = {
        photos_ws.cell(row=row, column=3).value
        for row in range(2, photos_ws.max_row + 1)
        if photos_ws.cell(row=row, column=3).value
    }
    assert photo_urls[0] in sources
    assert len(getattr(photos_ws, "_images", [])) >= len(photo_urls)


def test_ranger_only_sees_own_inspections(seeded_app: TruckInspectionApp, ranger: User, supervisor: User, truck) -> None:
    other = seeded_app.auth.register_user(
        name="Test Ranger",
        email="test@email.com",
        password="password123",
        role=UserRole.RANGER,
        ranger_number="RN-9001",
        security_responses=[
            ("Favorite trail?", "Canyon"),
            ("First ranger station?", "Pinecrest"),
            ("Go-to field snack?", "Jerky"),
        ],
    )
    seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.QUICK,
        responses=_quick_responses(),
        photo_urls=_photos(),
    )
    seeded_app.submit_inspection(
        user=other,
        truck=truck,
        inspection_type=InspectionType.QUICK,
        responses=_quick_responses(),
        photo_urls=_photos(),
    )
    ranger_inspections = seeded_app.list_inspections(requester=ranger)
    assert len(ranger_inspections) == 1
    supervisor_inspections = seeded_app.list_inspections(requester=supervisor)
    assert len(supervisor_inspections) == 2


def test_seeded_truck_identifiers(seeded_app: TruckInspectionApp) -> None:
    identifiers = {truck.identifier for truck in seeded_app.list_trucks()}
    assert identifiers == {
        "427",
        "P0101",
        "P0103",
        "P0106",
        "SM88",
        "T1",
        "T2",
        "T3",
    }


def test_checkout_and_return_flow(seeded_app: TruckInspectionApp, ranger: User, truck) -> None:
    start_inspection = seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.QUICK,
        responses={
            "exterior_clean": True,
            "interior_clean": True,
            "seatbelts_functioning": True,
            "tire_inflation": True,
            "fuel_level": "80",
            "odometer_miles": 1000,
        },
        photo_urls=["start1.jpg", "start2.jpg", "start3.jpg", "start4.jpg"],
    )

    assignment = seeded_app.checkout_truck(ranger=ranger, truck=truck, inspection=start_inspection)
    assert assignment.start_miles == 1000

    available_ids = {t.id for t in seeded_app.list_available_trucks()}
    assert truck.id not in available_ids

    end_inspection = seeded_app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.RETURN,
        responses={
            "odometer_miles": 1012,
            "return_notes": "All clear",
        },
        photo_urls=[],
    )

    completed = seeded_app.return_truck(assignment_id=assignment.id, ranger=ranger, inspection=end_inspection)
    assert completed.end_miles == 1012
    assert completed.returned_at is not None

    available_ids = {t.id for t in seeded_app.list_available_trucks()}
    assert truck.id in available_ids


def test_register_user_allowlist(app: TruckInspectionApp) -> None:
    user = app.auth.register_user(
        name="Angel Rodriguez",
        email="angel.rodriguezii@denvergov.org",
        password="securepass",
        role=UserRole.RANGER,
        ranger_number="RN-3001",
        security_responses=[
            ("Favorite lookout?", "Sunrise Point"),
            ("First ranger partner?", "Luis"),
            ("Best campsite?", "Bear Creek"),
        ],
    )
    assert user.email == "angel.rodriguezii@denvergov.org"
    assert user.ranger_number == "RN-3001"


def test_register_user_disallowed(app: TruckInspectionApp) -> None:
    with pytest.raises(ValueError):
        app.auth.register_user(
            name="Unauthorized",
            email="unauthorized@example.com",
            password="password123",
            role=UserRole.RANGER,
            ranger_number="RN-9999",
            security_responses=[
                ("Q1", "A1"),
                ("Q2", "A2"),
                ("Q3", "A3"),
            ],
        )


def test_update_password(app: TruckInspectionApp) -> None:
    user = app.auth.register_user(
        name="Test User",
        email="test@email.com",
        password="initialpass",
        ranger_number="RN-4001",
        security_responses=[
            ("Favorite overlook?", "High Point"),
            ("First badge number?", "42"),
            ("Preferred route?", "River Path"),
        ],
    )
    updated = app.auth.update_password(
        "test@email.com",
        "newpass",
        ["High Point", "42", "River Path"],
    )
    assert updated.id == user.id
    assert updated.ranger_number == "RN-4001"
    token = app.auth.authenticate("test@email.com", "newpass")
    assert token is not None


def test_reserve_truck(app: TruckInspectionApp) -> None:
    ranger = app.auth.register_user(
        name="Reservation Ranger",
        email="reserve@example.com",
        password="securepass",
        ranger_number="RN-5001",
        security_responses=[
            ("Go-to campsite?", "Lakeview"),
            ("Favorite creek?", "Pine"),
            ("Backup snack?", "Granola"),
        ],
    )
    supervisor = app.auth.register_user(
        name="Supervisor Reserve",
        email="super.reserve@example.com",
        password="superpass",
        role=UserRole.SUPERVISOR,
        ranger_number="RN-6001",
        security_responses=[
            ("Join year?", "2010"),
            ("First district?", "North"),
            ("Radio code?", "Bravo"),
        ],
    )
    truck = app.list_trucks()[0]

    default_reservation = app.reserve_truck(requester=ranger, truck=truck, note="")
    digits = "".join(ch for ch in (ranger.ranger_number or "") if ch.isdigit())
    suffix = digits[-2:] if len(digits) >= 2 else digits or "--"
    assert default_reservation.note == f"Reserved by Ranger {suffix}"

    reservation = app.update_reservation_note(requester=ranger, truck=truck, note="Hold for patrol")
    assert reservation.truck_id == truck.id
    assert reservation.user_id == ranger.id
    assert reservation.note == "Hold for patrol"

    with pytest.raises(ValueError):
        app.update_reservation_note(requester=ranger, truck=truck, note="x" * 81)

    other = app.auth.register_user(
        name="Other Ranger",
        email="other.reserve@example.com",
        password="pass1234",
        ranger_number="RN-5002",
        security_responses=[
            ("Fav park?", "Mesa"),
            ("Trail?", "North Loop"),
            ("Snack?", "Apple"),
        ],
    )
    with pytest.raises(ValueError):
        app.reserve_truck(requester=other, truck=truck, note="My turn")

    with pytest.raises(PermissionError):
        app.cancel_reservation(requester=supervisor, truck=truck)

    app.cancel_reservation(requester=ranger, truck=truck)
    assert app.database.get_reservation_for_truck(truck.id) is None

    reservation_again = app.reserve_truck(requester=ranger, truck=truck, note="")
    assert reservation_again.note == f"Reserved by Ranger {suffix}"
    inspection = app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.QUICK,
        responses={"odometer_miles": 200, "truck_clean": "yes"},
        photo_urls=[],
    )
    assignment = app.checkout_truck(ranger=ranger, truck=truck, inspection=inspection)
    assert assignment.truck_id == truck.id
    assert app.database.get_reservation_for_truck(truck.id) is None


def test_supervisor_cannot_return_other_assignment(app: TruckInspectionApp) -> None:
    supervisor = app.auth.register_user(
        name="Supervisor Check",
        email="super.check@example.com",
        password="adminpass",
        role=UserRole.SUPERVISOR,
        ranger_number="RN-6101",
        security_responses=[
            ("Fav lookout?", "North Rim"),
            ("Badge?", "77"),
            ("Snack?", "Peanuts"),
        ],
    )
    ranger = app.auth.register_user(
        name="Return Owner",
        email="owner@example.com",
        password="ownerpass",
        ranger_number="RN-5201",
        security_responses=[
            ("Camp?", "Lakeside"),
            ("Trail?", "South"),
            ("Snack?", "Gorp"),
        ],
    )
    truck = app.list_trucks()[0]

    checkout = app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.QUICK,
        responses={"odometer_miles": 1500, "truck_clean": "yes"},
        photo_urls=[],
    )
    assignment = app.checkout_truck(ranger=ranger, truck=truck, inspection=checkout)

    return_inspection = app.submit_inspection(
        user=ranger,
        truck=truck,
        inspection_type=InspectionType.RETURN,
        responses={"odometer_miles": 1510, "return_notes": "Ready"},
        photo_urls=[],
    )

    with pytest.raises(PermissionError):
        app.return_truck(assignment_id=assignment.id, ranger=supervisor, inspection=return_inspection)

    completed = app.return_truck(assignment_id=assignment.id, ranger=ranger, inspection=return_inspection)
    assert completed.returned_at is not None
